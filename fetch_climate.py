# -*- coding: utf-8 -*-
"""
Daily IPMA climate data fetcher for GitHub Actions.
Appends the last 24h of hourly observations for the nearest
station to Guimarães to a master CSV file.
Runs headless — no pyrevit, no GUI, no ctypes.

OUTPUT: data/climate_guimaraes.csv  (UTF-8, header on row 1)
  - On the FIRST run, if the CSV does not exist yet but the old
    climate_guimaraes.xlsx does, all history is migrated from the
    xlsx into the CSV so nothing is lost.
  - After that, new hours are appended to the CSV daily.
  - The old xlsx is never written again (frozen backup).
"""

import os
import csv
import json
import math
import datetime
import urllib.request

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
TARGET_LAT  = 41.4434
TARGET_LON  = -8.2938
TARGET_NAME = "Guimarães"

# Fallback station for precipitation (Guimarães station 1210625 has no precip sensor)
# 1210881 is the nearest station that reliably reports precAcumulada
PRECIP_FALLBACK_ID = None  # will be found automatically at runtime

IPMA_STATIONS_URL = "https://api.ipma.pt/open-data/observation/meteorology/stations/stations.json"
IPMA_OBS_URL      = "https://api.ipma.pt/open-data/observation/meteorology/stations/observations.json"

DATA_DIR   = "data"
CSV_FILE   = os.path.join(DATA_DIR, "climate_guimaraes.csv")
XLSX_FILE  = os.path.join(DATA_DIR, "climate_guimaraes.xlsx")  # old file — read once for migration only

os.makedirs(DATA_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def fetch_json(url):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/126.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
        }
    )
    response = urllib.request.urlopen(req, timeout=15)
    return json.loads(response.read().decode("utf-8"))

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = (math.sin(d_lat/2)**2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(d_lon/2)**2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

NO_DATA = -99.0
def clean(val, decimals=1):
    if val is None:
        return None
    try:
        fval = float(val)
    except (ValueError, TypeError):
        return None
    return None if fval == NO_DATA else round(fval, decimals)

WIND_DIR = {
    0: "Calm", 1: "N", 2: "NE", 3: "E", 4: "SE",
    5: "S",    6: "SW", 7: "W", 8: "NW", 9: "N"
}

COLUMNS = [
    "Timestamp (UTC)", "Station",
    "Temperature (°C)", "Humidity (%)", "Precipitation (mm)",
    "Wind Speed (km/h)", "Wind Speed (m/s)", "Wind Direction",
    "Pressure (hPa)", "Radiation (W/m²)"
]


# ---------------------------------------------------------------------------
# STEP 1 — Find nearest station
# ---------------------------------------------------------------------------
print("Finding nearest station to {}...".format(TARGET_NAME))
stations = fetch_json(IPMA_STATIONS_URL)

best_id   = None
best_name = None
best_dist = float("inf")

for feature in stations:
    props  = feature.get("properties", {})
    coords = feature.get("geometry", {}).get("coordinates", [])
    if not coords or len(coords) < 2:
        continue
    dist = haversine(TARGET_LAT, TARGET_LON, float(coords[1]), float(coords[0]))
    if dist < best_dist:
        best_dist = dist
        best_id   = str(props.get("idEstacao", ""))
        best_name = props.get("localEstacao", best_id)

print("Nearest station: {} (ID: {}) — {:.1f} km".format(best_name, best_id, best_dist))

# ---------------------------------------------------------------------------
# STEP 1b — Find nearest station with valid precipitation sensor
# ---------------------------------------------------------------------------
# Build a lookup of all stations by id
station_coords = {}
for feature in stations:
    props  = feature.get("properties", {})
    coords = feature.get("geometry", {}).get("coordinates", [])
    if not coords or len(coords) < 2: continue
    sid = str(props.get("idEstacao", ""))
    station_coords[sid] = (float(coords[1]), float(coords[0]), props.get("localEstacao", sid))

# ---------------------------------------------------------------------------
# STEP 2 — Fetch observations
# ---------------------------------------------------------------------------
print("Fetching observations...")
obs_data = fetch_json(IPMA_OBS_URL)

# Find nearest station with valid precip data from the latest timestamp
all_timestamps = sorted(obs_data.keys())
latest_ts = all_timestamps[-1]
precip_station_id   = best_id   # default to main station
precip_station_name = best_name
precip_best_dist    = float("inf")

for sid, sdata in obs_data[latest_ts].items():
    if sdata is None: continue
    prec = sdata.get("precAcumulada")
    if prec is None: continue
    try:
        if float(prec) == NO_DATA: continue
    except (ValueError, TypeError):
        continue
    if sid in station_coords:
        slat, slon, sname = station_coords[sid]
        dist = haversine(TARGET_LAT, TARGET_LON, slat, slon)
        if dist < precip_best_dist:
            precip_best_dist    = dist
            precip_station_id   = sid
            precip_station_name = sname

if precip_station_id != best_id:
    print("Precipitation from nearest reporting station: {} (ID: {}) — {:.1f} km".format(
        precip_station_name, precip_station_id, precip_best_dist
    ))
else:
    print("Main station reports precipitation — no fallback needed.")

new_records = []
missing_hours = []

for ts in all_timestamps:
    sd = obs_data[ts].get(best_id)
    if sd is None:
        missing_hours.append(ts)
        continue
    new_records.append({
        "Timestamp (UTC)":    ts,
        "Station":            best_name,
        "Temperature (°C)":   clean(sd.get("temperatura")),
        "Humidity (%)":       clean(sd.get("humidade")),
        "Precipitation (mm)": clean(obs_data[ts].get(precip_station_id, {}).get("precAcumulada") if obs_data[ts].get(precip_station_id) else None),
        "Wind Speed (km/h)":  clean(sd.get("intensidadeVentoKM")),
        "Wind Speed (m/s)":   clean(sd.get("intensidadeVento")),
        "Wind Direction":     WIND_DIR.get(sd.get("idDireccVento"), ""),
        "Pressure (hPa)":     clean(sd.get("pressao")),
        "Radiation (W/m²)":   clean(sd.get("radiacao")),
    })

print("{} records fetched.".format(len(new_records)))
if missing_hours:
    print("WARNING — {} hours with no station data (station gap): {}".format(
        len(missing_hours), ", ".join(missing_hours)
    ))

if not new_records:
    print("No data returned — skipping Excel update.")
    exit(0)

# ---------------------------------------------------------------------------
# STEP 3 - Write to CSV  (with one-time migration from the old xlsx)
# ---------------------------------------------------------------------------

def _ts_str(v):
    """Normalise a timestamp cell to the IPMA ISO string format."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%dT%H:%M")
    return str(v)


def migrate_xlsx_to_csv(xlsx_path, csv_path):
    """
    One-time migration: read all history from the old xlsx (header on row 3,
    data from row 4) and write it into a fresh CSV with the header on row 1.
    Uses openpyxl, which is available on the GitHub Actions runner.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Hourly Data"]
    n = 0
    # utf-8-sig writes a BOM so Excel opens the degree/accent characters correctly
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        for r in range(4, ws.max_row + 1):
            vals = [ws.cell(row=r, column=ci).value for ci in range(1, len(COLUMNS) + 1)]
            if vals[0] is None:           # no timestamp -> not a data row
                continue
            vals[0] = _ts_str(vals[0])    # normalise the timestamp
            writer.writerow(["" if v is None else v for v in vals])
            n += 1
    wb.close()
    return n


def read_existing_timestamps(csv_path):
    """Collect timestamps already in the CSV, to avoid duplicates."""
    existing = set()
    if not os.path.exists(csv_path):
        return existing
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        ts_idx = 0
        if header:
            for i, h in enumerate(header):
                if "timestamp" in str(h).lower():
                    ts_idx = i
                    break
        for row in reader:
            if row and len(row) > ts_idx and str(row[ts_idx]).strip():
                existing.add(str(row[ts_idx]).strip())
    return existing


def row_values(record):
    """Turn a record dict into a list matching COLUMNS order."""
    out = []
    for col in COLUMNS:
        v = record.get(col)
        out.append("" if v is None else v)
    return out


# --- 3a. Seed the CSV on first run -----------------------------------------
if not os.path.exists(CSV_FILE):
    if os.path.exists(XLSX_FILE):
        moved = migrate_xlsx_to_csv(XLSX_FILE, CSV_FILE)
        print("First run: migrated {} historical rows from xlsx into the new CSV.".format(moved))
    else:
        with open(CSV_FILE, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow(COLUMNS)
        print("First run: created a new CSV with header (no xlsx found to migrate).")

# --- 3b. Append new records + gap rows -------------------------------------
existing_ts = read_existing_timestamps(CSV_FILE)

appended = skipped = gaps_added = 0
# append mode uses plain utf-8 (no BOM) so we don't insert a BOM mid-file
with open(CSV_FILE, "a", encoding="utf-8", newline="") as f:
    writer = csv.writer(f)

    for record in new_records:
        ts = record["Timestamp (UTC)"]
        if ts in existing_ts:
            skipped += 1
            continue
        writer.writerow(row_values(record))
        existing_ts.add(ts)
        appended += 1

    for ts in missing_hours:
        if ts in existing_ts:
            continue
        writer.writerow(row_values({
            "Timestamp (UTC)": ts,
            "Station": best_name,
            "Temperature (°C)": "NO DATA",
        }))
        existing_ts.add(ts)
        gaps_added += 1

print("{} rows appended, {} duplicates skipped, {} gap rows added.".format(
    appended, skipped, gaps_added))
print("Saved to {}".format(CSV_FILE))
print("Done!")
